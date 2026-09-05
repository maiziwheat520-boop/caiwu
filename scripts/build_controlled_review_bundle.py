"""Build the local, transient source bundle for the 2026-05 review cutover.

This command intentionally emits only counts and the manifest digest.  Row
values stay inside the private bundle and are never printed to the terminal.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import os
import shutil
import stat
from datetime import UTC, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID, uuid5

from ledgerbridge.controlled_import import SOURCE_MANIFEST_SCHEMA, SourceManifest

_NAMESPACE = UUID("b2f82a31-26cf-4b43-a6a7-8e90339ab468")
_EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PHOTO_CELLS = (
    ("B4", "薇旭美团", 1),
    ("B5", "薇旭美团", 2),
    ("B6", "薇旭美团", 3),
    ("B7", "薇旭美团", 4),
    ("C4", "薇旭携程", 1),
    ("C5", "薇旭携程", 2),
    ("C6", "薇旭携程", 3),
    ("C7", "薇旭携程", 4),
    ("B13", "景怡美团", 1),
    ("B14", "景怡美团", 2),
    ("B15", "景怡美团", 3),
    ("B16", "景怡美团", 4),
)


class BundleBuildError(RuntimeError):
    pass


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--combined-workbook", type=Path, required=True)
    parser.add_argument("--manual-review-workbook", type=Path, required=True)
    parser.add_argument("--photo-directory", type=Path, required=True)
    parser.add_argument("--ocr-observations", type=Path)
    parser.add_argument("--output-directory", type=Path, required=True)
    args = parser.parse_args()
    manifest_path, manifest = build_bundle(
        combined_workbook=args.combined_workbook.resolve(),
        manual_review_workbook=args.manual_review_workbook.resolve(),
        photo_directory=args.photo_directory.resolve(),
        output_directory=args.output_directory.resolve(),
        ocr_observations=args.ocr_observations.resolve() if args.ocr_observations else None,
    )
    digest = hashlib.sha256(manifest_path.read_bytes()).hexdigest()
    print(
        "CONTROLLED_REVIEW_BUNDLE_OK "
        f"evidence={len(manifest.evidence)} candidates={len(manifest.candidates)} "
        f"manifest_sha256={digest}"
    )
    return 0


def build_bundle(
    *,
    combined_workbook: Path,
    manual_review_workbook: Path,
    photo_directory: Path,
    output_directory: Path,
    ocr_observations: Path | None = None,
) -> tuple[Path, SourceManifest]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise BundleBuildError("openpyxl is required to build the controlled bundle") from exc
    _require_regular_file(combined_workbook)
    _require_regular_file(manual_review_workbook)
    if not photo_directory.is_dir():
        raise BundleBuildError("photo directory is unavailable")
    photos = sorted(
        (
            path
            for path in photo_directory.iterdir()
            if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png"}
        ),
        key=lambda path: path.name.casefold(),
    )
    if len(photos) != 5:
        raise BundleBuildError("exactly five source photos are required")
    if output_directory.exists():
        raise BundleBuildError("output directory already exists")
    output_directory.mkdir(mode=0o700, parents=False)

    evidence: list[dict[str, object]] = []
    photo_evidence_refs: list[UUID] = []
    photo_ref_by_source_name: dict[str, UUID] = {}
    source_paths = [*photos, combined_workbook, manual_review_workbook]
    if len({_digest(path) for path in photos}) != len(photos):
        raise BundleBuildError("source photos must have unique digests")
    normalized_sources: list[tuple[Path, str, str]] = []
    for index, source in enumerate(photos, start=1):
        suffix = ".jpg" if source.suffix.lower() in {".jpg", ".jpeg"} else ".png"
        normalized_sources.append((source, f"photo-{index:02d}{suffix}", _media_type(source)))
    normalized_sources.extend(
        (
            (combined_workbook, "combined-review.xlsx", _EXCEL_MEDIA_TYPE),
            (manual_review_workbook, "boc-manual-review.xlsx", _EXCEL_MEDIA_TYPE),
        )
    )
    for source, normalized_name, media_type in normalized_sources:
        digest = _digest(source)
        evidence_ref = uuid5(_NAMESPACE, f"evidence:{normalized_name}:{digest}")
        shutil.copy2(source, output_directory / normalized_name)
        evidence.append(
            {
                "evidence_ref": evidence_ref,
                "source_file": normalized_name,
                "display_name": normalized_name,
                "declared_media_type": media_type,
                "plaintext_sha256": digest,
                "plaintext_size": source.stat().st_size,
            }
        )
        if normalized_name.startswith("photo-"):
            photo_evidence_refs.append(evidence_ref)
            photo_ref_by_source_name[source.name] = evidence_ref
    combined_ref = UUID(str(evidence[-2]["evidence_ref"]))
    manual_review_ref = UUID(str(evidence[-1]["evidence_ref"]))

    workbook = load_workbook(combined_workbook, read_only=True, data_only=True)
    try:
        if "26.5" not in workbook.sheetnames:
            raise BundleBuildError("combined workbook is missing the 26.5 sheet")
        photo_sheet = workbook["26.5"]
        email_sheets = [sheet for sheet in workbook.worksheets if "邮箱待复核" in sheet.title]
        if len(email_sheets) != 1:
            raise BundleBuildError("combined workbook must contain one email review sheet")
        email_sheet = email_sheets[0]
        if ocr_observations is None:
            candidates = _photo_candidates(
                photo_sheet,
                evidence_refs=tuple((*photo_evidence_refs, combined_ref)),
            )
        else:
            candidates = _ocr_photo_candidates(
                ocr_observations,
                evidence_ref_by_source_name=photo_ref_by_source_name,
            )
        candidates.extend(
            _boc_candidates(
                email_sheet,
                evidence_refs=(combined_ref, manual_review_ref),
            )
        )
    finally:
        workbook.close()
    if ocr_observations is None and len(candidates) != 54:
        raise BundleBuildError("legacy controlled review bundle must contain exactly 54 candidates")
    if ocr_observations is not None and len(candidates) <= 42:
        raise BundleBuildError("OCR review bundle did not produce any review-ready photo bills")

    input_fingerprint = ":".join(_digest(path) for path in source_paths)
    batch_ref = uuid5(_NAMESPACE, f"batch:2026-05:{input_fingerprint}")
    generated_at = datetime.fromtimestamp(
        max(path.stat().st_mtime for path in source_paths),
        tz=UTC,
    )
    entity_ref = uuid5(_NAMESPACE, "entity:controlled-reconciliation")
    business_unit_ref = uuid5(_NAMESPACE, "business-unit:2026-05-controlled-review")
    payload: dict[str, Any] = {
        "schema_version": SOURCE_MANIFEST_SCHEMA,
        "batch_ref": batch_ref,
        "generated_at": generated_at,
        "source_description": (
            "Controlled review import of five original hotel screenshots and 42 BOC rows "
            "derived from the prior review workbooks; the original mail attachment is not "
            "bundled and the BOC rows remain derived review evidence."
        ),
        "entity": {
            "entity_ref": entity_ref,
            "name": "LedgerBridge controlled reconciliation",
        },
        "business_unit": {
            "business_unit_ref": business_unit_ref,
            "ref": "review-2026-05",
            "label": "2026年5月对账复核",
        },
        "categories": (
            {
                "category_ref": uuid5(_NAMESPACE, "category:photo-reconciliation"),
                "code": "PHOTO_RECONCILIATION",
                "label": "照片渠道对账",
            },
            {
                "category_ref": uuid5(_NAMESPACE, "category:boc-transaction-review"),
                "code": "BOC_TRANSACTION_REVIEW",
                "label": "中行交易复核",
            },
        ),
        "evidence": tuple(evidence),
        "candidates": tuple(candidates),
    }
    manifest = SourceManifest.model_validate(payload, strict=True)
    manifest_path = output_directory / "source-manifest.json"
    _write_private_json(manifest_path, manifest.model_dump(mode="json"))
    if ocr_observations is not None:
        try:
            from scripts.build_hotel_payout_cutover import (
                build_hotel_payout_cutover_manifest,
            )
        except ModuleNotFoundError:  # direct ``python scripts/...`` execution
            from build_hotel_payout_cutover import (  # type: ignore[no-redef,import-not-found]
                build_hotel_payout_cutover_manifest,
            )

        build_hotel_payout_cutover_manifest(
            source_manifest_path=manifest_path.resolve(),
            combined_workbook=combined_workbook,
            ocr_observations=ocr_observations,
            output_path=(output_directory / "hotel-payout-cutover.json").resolve(),
        )
    return manifest_path, manifest


def _photo_candidates(sheet: Any, *, evidence_refs: tuple[UUID, ...]) -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    for coordinate, group, week in _PHOTO_CELLS:
        amount_minor = _money_minor(sheet[coordinate].value)
        stable = f"photo:2026-05:{coordinate}:{group}:{week}"
        candidate_ref = uuid5(_NAMESPACE, f"candidate:{stable}")
        candidates.append(
            {
                "candidate_ref": candidate_ref,
                "operation_id": uuid5(_NAMESPACE, f"operation:{stable}"),
                "ingest_channel": "CONTROLLED_UPLOAD",
                "source_system": "hotel_photo_reconciliation",
                "source_event_ref": uuid5(_NAMESPACE, f"source-event:{stable}"),
                "display_label": "Hotel photo reconciliation",
                "category_code": "PHOTO_RECONCILIATION",
                "amount_minor": amount_minor,
                "accounting_month": "2026-05",
                "summary": f"照片待复核: {group}第{week}周 ({coordinate})",
                "confidence_basis_points": 8500,
                "evidence_refs": evidence_refs,
            }
        )
    return candidates


def _ocr_photo_candidates(
    ocr_observations: Path,
    *,
    evidence_ref_by_source_name: dict[str, UUID],
) -> list[dict[str, object]]:
    try:
        metadata = ocr_observations.lstat()
    except OSError as exc:
        raise BundleBuildError("OCR observations are unavailable") from exc
    if stat.S_ISLNK(metadata.st_mode) or not stat.S_ISREG(metadata.st_mode):
        raise BundleBuildError("OCR observations must be a regular file")
    if metadata.st_size > 16 * 1024 * 1024:
        raise BundleBuildError("OCR observations exceed the size limit")
    try:
        raw = ocr_observations.read_bytes()
    except OSError as exc:
        raise BundleBuildError("OCR observations are unavailable") from exc
    try:
        payload = json.loads(raw)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise BundleBuildError("OCR observations are invalid JSON") from exc
    if not isinstance(payload, dict) or payload.get("schema_version") != "ledgerbridge.bill-ocr.v1":
        raise BundleBuildError("OCR observation schema is unsupported")
    results = payload.get("results")
    if not isinstance(results, list):
        raise BundleBuildError("OCR observation results are invalid")
    candidates: list[dict[str, object]] = []
    for result in results:
        if not isinstance(result, dict):
            raise BundleBuildError("OCR observation result is invalid")
        source_name = result.get("source_name")
        bills = result.get("bills")
        if not isinstance(source_name, str) or not isinstance(bills, list):
            raise BundleBuildError("OCR observation source is invalid")
        evidence_ref = evidence_ref_by_source_name.get(source_name)
        if evidence_ref is None:
            raise BundleBuildError("OCR observation does not match a source image")
        for bill in bills:
            if not isinstance(bill, dict):
                raise BundleBuildError("OCR bill observation is invalid")
            blockers = bill.get("blockers")
            if not isinstance(blockers, list):
                raise BundleBuildError("OCR bill blockers are invalid")
            if blockers:
                continue
            bill_id = bill.get("bill_id")
            period_start = bill.get("period_start")
            period_end = bill.get("period_end")
            amount_minor = bill.get("amount_minor")
            confidence = bill.get("confidence_basis_points")
            source_kind = bill.get("source_kind")
            if (
                not isinstance(bill_id, str)
                or not isinstance(period_start, str)
                or not isinstance(period_end, str)
                or isinstance(amount_minor, bool)
                or not isinstance(amount_minor, int)
                or isinstance(confidence, bool)
                or not isinstance(confidence, int)
                or not isinstance(source_kind, str)
            ):
                raise BundleBuildError("review-ready OCR bill fields are incomplete")
            if not (period_start.startswith("2026-05") or period_end.startswith("2026-05")):
                continue
            stable = f"ocr:{source_name}:{bill_id}:{period_start}:{period_end}"
            candidates.append(
                {
                    "candidate_ref": uuid5(_NAMESPACE, f"candidate:{stable}"),
                    "operation_id": uuid5(_NAMESPACE, f"operation:{stable}"),
                    "ingest_channel": "CONTROLLED_UPLOAD",
                    "source_system": "hotel_bill_ocr",
                    "source_event_ref": uuid5(_NAMESPACE, f"source-event:{stable}"),
                    "display_label": "OCR bill pending review",
                    "category_code": "PHOTO_RECONCILIATION",
                    "amount_minor": amount_minor,
                    "accounting_month": "2026-05",
                    "summary": f"OCR账单待复核: {source_kind} {bill_id}",
                    "confidence_basis_points": confidence,
                    "evidence_refs": (evidence_ref,),
                }
            )
    return candidates


def _boc_candidates(sheet: Any, *, evidence_refs: tuple[UUID, ...]) -> list[dict[str, object]]:
    headers = {
        str(sheet.cell(4, column).value).strip(): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(4, column).value is not None
    }
    required = {"清单ID", "复核状态", "银行", "交易时间", "金额(元)", "附件SHA256", "机器数值校验"}
    if not required <= set(headers):
        raise BundleBuildError("email review sheet headers are incomplete")
    candidates: list[dict[str, object]] = []
    for row in range(5, sheet.max_row + 1):
        item_id = str(sheet.cell(row, headers["清单ID"]).value or "").strip()
        if not item_id:
            continue
        status = str(sheet.cell(row, headers["复核状态"]).value or "").strip()
        bank = str(sheet.cell(row, headers["银行"]).value or "").strip()
        machine = str(sheet.cell(row, headers["机器数值校验"]).value or "").strip()
        attachment_sha = str(sheet.cell(row, headers["附件SHA256"]).value or "").strip().lower()
        if status != "待复核" or bank != "BOC" or machine != "通过":
            raise BundleBuildError("email review row is outside the authorized pending set")
        if len(attachment_sha) != 64 or any(ch not in "0123456789abcdef" for ch in attachment_sha):
            raise BundleBuildError("email review row has an invalid attachment digest")
        transaction_time = sheet.cell(row, headers["交易时间"]).value
        if isinstance(transaction_time, datetime):
            accounting_month = transaction_time.strftime("%Y-%m")
        else:
            accounting_month = str(transaction_time or "")[:7]
        if accounting_month != "2026-05":
            raise BundleBuildError("email review row is outside 2026-05")
        amount_minor = _money_minor(sheet.cell(row, headers["金额(元)"]).value)
        stable = f"boc-derived:{item_id}:{attachment_sha}"
        candidate_ref = uuid5(_NAMESPACE, f"candidate:{stable}")
        candidates.append(
            {
                "candidate_ref": candidate_ref,
                "operation_id": uuid5(_NAMESPACE, f"operation:{stable}"),
                "ingest_channel": "OUTLOOK",
                "source_system": "boc_mail_derived_review",
                "source_event_ref": uuid5(_NAMESPACE, f"source-event:{stable}"),
                "display_label": "BOC email derived review row",
                "category_code": "BOC_TRANSACTION_REVIEW",
                "amount_minor": amount_minor,
                "accounting_month": accounting_month,
                "summary": f"中行邮箱账单待复核: {item_id}",
                "confidence_basis_points": 7000,
                "evidence_refs": evidence_refs,
            }
        )
    if len(candidates) != 42:
        raise BundleBuildError("email review sheet must contain exactly 42 authorized rows")
    return candidates


def _money_minor(value: object) -> int:
    if isinstance(value, bool) or value is None:
        raise BundleBuildError("candidate amount is missing")
    try:
        decimal = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise BundleBuildError("candidate amount is invalid") from exc
    minor = (decimal * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    if decimal != minor / 100:
        raise BundleBuildError("candidate amount has more than two decimal places")
    return int(minor)


def _media_type(path: Path) -> str:
    value, _ = mimetypes.guess_type(path.name)
    if value not in {"image/jpeg", "image/png"}:
        raise BundleBuildError("source photo media type is unsupported")
    return value


def _digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            hasher.update(chunk)
    return hasher.hexdigest()


def _require_regular_file(path: Path) -> None:
    try:
        metadata = path.lstat()
    except OSError as exc:
        raise BundleBuildError("source workbook is unavailable") from exc
    if stat.S_ISLNK(metadata.st_mode) or not stat.S_ISREG(metadata.st_mode):
        raise BundleBuildError("source workbook must be a regular file")


def _write_private_json(path: Path, payload: dict[str, Any]) -> None:
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("ascii")
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0)
    descriptor = os.open(path, flags, 0o600)
    try:
        written = 0
        while written < len(encoded):
            count = os.write(descriptor, encoded[written:])
            if count <= 0:
                raise BundleBuildError("manifest write made no progress")
            written += count
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


if __name__ == "__main__":
    raise SystemExit(main())
