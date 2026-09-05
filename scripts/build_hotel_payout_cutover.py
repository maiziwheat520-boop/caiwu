"""Build a private OCR replacement and hotel-bank evidence-link manifest."""

from __future__ import annotations

import argparse
import hashlib
import json
import stat
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import UUID

from ledgerbridge.controlled_import import SourceManifest, load_source_manifest
from ledgerbridge.hotel_payout_cutover import (
    HOTEL_PAYOUT_CUTOVER_SCHEMA,
    CandidateEvidenceLink,
    HotelMatchBasis,
    HotelPayoutCutoverManifest,
    HotelReplacement,
    write_private_cutover_manifest,
)

try:
    from scripts.build_controlled_review_bundle import (
        _NAMESPACE,
        _boc_candidates,
        _photo_candidates,
    )
except ModuleNotFoundError:  # direct ``python scripts/...`` execution
    from build_controlled_review_bundle import (  # type: ignore[no-redef,import-not-found]
        _NAMESPACE,
        _boc_candidates,
        _photo_candidates,
    )

_CTRIP_TERMS = ("携程", "赫程", "旅行社")
_MEITUAN_TERMS = ("美团", "钱袋宝", "宝支付")


class HotelPayoutBuildError(RuntimeError):
    pass


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-manifest", type=Path, required=True)
    parser.add_argument("--combined-workbook", type=Path, required=True)
    parser.add_argument("--ocr-observations", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    manifest = build_hotel_payout_cutover_manifest(
        source_manifest_path=args.source_manifest.resolve(),
        combined_workbook=args.combined_workbook.resolve(),
        ocr_observations=args.ocr_observations.resolve(),
        output_path=args.output.resolve(),
    )
    print(
        "HOTEL_PAYOUT_CUTOVER_OK "
        f"ocr={len(manifest.ocr_candidate_refs)} "
        f"replaced={len(manifest.replacements)} links={len(manifest.evidence_links)} "
        f"unmatched={len(manifest.ocr_candidate_refs) - len(manifest.evidence_links)}"
    )
    return 0


def build_hotel_payout_cutover_manifest(
    *,
    source_manifest_path: Path,
    combined_workbook: Path,
    ocr_observations: Path,
    output_path: Path,
) -> HotelPayoutCutoverManifest:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise HotelPayoutBuildError("openpyxl is required") from exc
    source, source_raw = load_source_manifest(source_manifest_path)
    observations = _load_ocr_observations(ocr_observations)
    if output_path.exists():
        raise HotelPayoutBuildError("cutover output already exists")
    if output_path.parent != source_manifest_path.parent:
        raise HotelPayoutBuildError("cutover manifest must stay beside the source manifest")

    workbook = load_workbook(combined_workbook, read_only=True, data_only=True)
    try:
        if "26.5" not in workbook.sheetnames:
            raise HotelPayoutBuildError("combined workbook is missing the 26.5 sheet")
        email_sheets = [sheet for sheet in workbook.worksheets if "邮箱待复核" in sheet.title]
        if len(email_sheets) != 1:
            raise HotelPayoutBuildError("combined workbook must contain one email review sheet")
        legacy_candidates = _photo_candidates(workbook["26.5"], evidence_refs=(UUID(int=1),))
        bank_candidates = _boc_candidates(email_sheets[0], evidence_refs=(UUID(int=1),))
        bank_rows = _bank_rows(email_sheets[0], bank_candidates)
    finally:
        workbook.close()

    source_by_ref = {item.candidate_ref: item for item in source.candidates}
    ocr_details = _ocr_details(observations, source)
    ocr_candidates = tuple(source_by_ref[item["candidate_ref"]] for item in ocr_details)
    replacements = _replacement_rows(legacy_candidates, ocr_candidates)
    links = _evidence_links(ocr_details, bank_rows)
    if len(links) != 8:
        raise HotelPayoutBuildError("expected eight unique hotel payout bank matches")
    source_digest = hashlib.sha256(source_raw).hexdigest()
    cutover = HotelPayoutCutoverManifest.model_validate(
        {
            "schema_version": HOTEL_PAYOUT_CUTOVER_SCHEMA,
            "cutover_ref": _uuid5(f"hotel-payout-cutover:{source_digest}"),
            "generated_at": source.generated_at,
            "source_manifest_sha256": source_digest,
            "entity_ref": source.entity.entity_ref,
            "business_unit_ref": source.business_unit.business_unit_ref,
            "ocr_candidate_refs": tuple(item.candidate_ref for item in ocr_candidates),
            "replacements": replacements,
            "evidence_links": links,
        },
        strict=True,
    )
    write_private_cutover_manifest(output_path, cutover)
    return cutover


def _load_ocr_observations(path: Path) -> dict[str, Any]:
    try:
        metadata = path.lstat()
    except OSError as exc:
        raise HotelPayoutBuildError("OCR observations are unavailable") from exc
    if stat.S_ISLNK(metadata.st_mode) or not stat.S_ISREG(metadata.st_mode):
        raise HotelPayoutBuildError("OCR observations must be a regular file")
    if metadata.st_size <= 0 or metadata.st_size > 16 * 1024 * 1024:
        raise HotelPayoutBuildError("OCR observations size is invalid")
    try:
        payload = json.loads(path.read_bytes())
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HotelPayoutBuildError("OCR observations are invalid JSON") from exc
    if not isinstance(payload, dict) or payload.get("schema_version") != "ledgerbridge.bill-ocr.v1":
        raise HotelPayoutBuildError("OCR observation schema is unsupported")
    return payload


def _ocr_details(payload: dict[str, Any], source: SourceManifest) -> tuple[dict[str, Any], ...]:
    source_refs = {
        item.candidate_ref for item in source.candidates if item.source_system == "hotel_bill_ocr"
    }
    details: list[dict[str, Any]] = []
    results = payload.get("results")
    if not isinstance(results, list):
        raise HotelPayoutBuildError("OCR observation results are invalid")
    for result in results:
        if not isinstance(result, dict):
            raise HotelPayoutBuildError("OCR observation result is invalid")
        source_name = result.get("source_name")
        bills = result.get("bills")
        if not isinstance(source_name, str) or not isinstance(bills, list):
            raise HotelPayoutBuildError("OCR observation source is invalid")
        for bill in bills:
            if not isinstance(bill, dict) or not isinstance(bill.get("blockers"), list):
                raise HotelPayoutBuildError("OCR bill is invalid")
            if bill["blockers"]:
                continue
            bill_id = bill.get("bill_id")
            period_start = bill.get("period_start")
            period_end = bill.get("period_end")
            platform = bill.get("source_kind")
            amount_minor = bill.get("amount_minor")
            if (
                not isinstance(bill_id, str)
                or not isinstance(period_start, str)
                or not isinstance(period_end, str)
                or not isinstance(platform, str)
                or not isinstance(amount_minor, int)
            ):
                raise HotelPayoutBuildError("review-ready OCR bill fields are incomplete")
            if not (period_start.startswith("2026-05") or period_end.startswith("2026-05")):
                continue
            stable = f"ocr:{source_name}:{bill_id}:{period_start}:{period_end}"
            candidate_ref = _uuid5(f"candidate:{stable}")
            if candidate_ref not in source_refs:
                raise HotelPayoutBuildError("OCR detail does not match source candidate")
            details.append(
                {
                    "candidate_ref": candidate_ref,
                    "amount_minor": amount_minor,
                    "platform": platform,
                    "period_start": period_start,
                    "period_end": period_end,
                }
            )
    if {item["candidate_ref"] for item in details} != source_refs:
        raise HotelPayoutBuildError("OCR source candidates and observations differ")
    return tuple(details)


def _bank_rows(sheet: Any, candidates: list[dict[str, object]]) -> tuple[dict[str, Any], ...]:
    headers = {
        str(sheet.cell(4, column).value).strip(): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(4, column).value is not None
    }
    by_id = {str(item["summary"]).rsplit(":", 1)[-1].strip(): item for item in candidates}
    rows: list[dict[str, Any]] = []
    for row_number in range(5, sheet.max_row + 1):
        item_id = str(sheet.cell(row_number, headers["清单ID"]).value or "").strip()
        if not item_id:
            continue
        transaction_time = sheet.cell(row_number, headers["交易时间"]).value
        if not isinstance(transaction_time, datetime):
            raise HotelPayoutBuildError("bank transaction time is invalid")
        candidate = by_id[item_id]
        text_value = " ".join(
            str(sheet.cell(row_number, column).value or "")
            for name, column in headers.items()
            if name not in {"附件SHA256", "机器数值校验"}
        )
        rows.append(
            {
                "candidate_ref": candidate["candidate_ref"],
                "amount_minor": candidate["amount_minor"],
                "date": transaction_time.date(),
                "item_id": item_id,
                "text": text_value,
            }
        )
    return tuple(rows)


def _replacement_rows(
    legacy_candidates: list[dict[str, object]],
    ocr_candidates: tuple[Any, ...],
) -> tuple[HotelReplacement, ...]:
    legacy_by_amount: dict[int, list[UUID]] = {}
    for item in legacy_candidates:
        amount = item["amount_minor"]
        if not isinstance(amount, int):
            raise HotelPayoutBuildError("legacy candidate amount is not a minor-unit integer")
        legacy_by_amount.setdefault(amount, []).append(UUID(str(item["candidate_ref"])))
    replacements: list[HotelReplacement] = []
    for candidate in ocr_candidates:
        matches = legacy_by_amount.get(candidate.amount_minor, [])
        if len(matches) == 1:
            replacements.append(
                HotelReplacement(
                    legacy_candidate_ref=matches[0],
                    ocr_candidate_ref=candidate.candidate_ref,
                    amount_minor=candidate.amount_minor,
                )
            )
        elif len(matches) > 1:
            raise HotelPayoutBuildError("legacy OCR replacement is ambiguous")
    if len(replacements) != 8:
        raise HotelPayoutBuildError("expected eight weekly summaries to be replaced")
    return tuple(replacements)


def _evidence_links(
    ocr_details: tuple[dict[str, Any], ...], bank_rows: tuple[dict[str, Any], ...]
) -> tuple[CandidateEvidenceLink, ...]:
    links: list[CandidateEvidenceLink] = []
    used_bank: set[UUID] = set()
    for detail in ocr_details:
        platform = detail["platform"]
        terms = _CTRIP_TERMS if platform == "CTRIP_EBOOKING" else _MEITUAN_TERMS
        period_end = datetime.fromisoformat(detail["period_end"]).date()
        matches = [
            row
            for row in bank_rows
            if row["amount_minor"] == detail["amount_minor"]
            and period_end <= row["date"] <= period_end + timedelta(days=7)
            and any(term in row["text"] for term in terms)
        ]
        if not matches:
            continue
        if len(matches) != 1:
            raise HotelPayoutBuildError("hotel payout bank match is ambiguous")
        bank = matches[0]
        bank_ref = UUID(str(bank["candidate_ref"]))
        if bank_ref in used_bank:
            raise HotelPayoutBuildError("bank credit was reused by multiple hotel payouts")
        used_bank.add(bank_ref)
        subject_ref = UUID(str(detail["candidate_ref"]))
        links.append(
            CandidateEvidenceLink(
                link_ref=_uuid5(f"hotel-payout-link:{subject_ref}:{bank_ref}"),
                subject_candidate_ref=subject_ref,
                evidence_candidate_ref=bank_ref,
                risk_code="HOTEL_PAYOUT_STATEMENT_REQUIRED",
                relation="SAME_ECONOMIC_TRANSACTION",
                amount_minor=detail["amount_minor"],
                currency="CNY",
                match_basis=HotelMatchBasis(
                    method="EXACT_AMOUNT_DATE_PLATFORM_ONE_TO_ONE",
                    platform=platform,
                    subject_period_start=detail["period_start"],
                    subject_period_end=detail["period_end"],
                    evidence_date=bank["date"],
                    evidence_transaction_ref=bank["item_id"],
                ),
            )
        )
    return tuple(links)


def _uuid5(value: str) -> UUID:
    from uuid import uuid5

    return uuid5(_NAMESPACE, value)


if __name__ == "__main__":
    raise SystemExit(main())
