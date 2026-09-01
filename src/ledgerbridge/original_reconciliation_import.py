"""Build controlled Candidate imports from a reviewed original-workbook mapping.

The workbook remains immutable evidence.  Only explicitly mapped, non-formula
cells become review Candidates; derived totals stay derived facts in the source
workbook and are never imported as transactions.
"""

from __future__ import annotations

import hashlib
import math
import re
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Literal
from uuid import UUID, uuid5

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, ConfigDict, Field, model_validator

from ledgerbridge.candidate_contract import JSON_SAFE_INTEGER
from ledgerbridge.controlled_import import (
    ImportBusinessUnit,
    ImportCandidate,
    ImportCategory,
    ImportEntity,
    SourceEvidence,
    SourceManifest,
)

_IMPORT_NAMESPACE = UUID("7dbe74df-7432-5cb4-9291-9789f55e693e")
_CELL = re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,5}$")
_SAFE_VERSION = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,99}$")


class OriginalReconciliationImportError(RuntimeError):
    """A workbook could not be converted without guessing financial facts."""


class _FrozenModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class OriginalReconciliationCellMapping(_FrozenModel):
    sheet_name: str = Field(min_length=1, max_length=100)
    cell: str = Field(pattern=_CELL.pattern)
    accounting_month: str = Field(pattern=r"^20[0-9]{2}-(0[1-9]|1[0-2])$")
    category_code: str = Field(pattern=r"^[A-Z][A-Z0-9_]{0,99}$")
    description: str = Field(min_length=1, max_length=400)
    sign_multiplier: Literal[-1, 1]


class OriginalReconciliationScopePlan(_FrozenModel):
    business_unit: ImportBusinessUnit
    categories: tuple[ImportCategory, ...] = Field(min_length=1)
    cells: tuple[OriginalReconciliationCellMapping, ...] = Field(min_length=1)

    @model_validator(mode="after")
    def mappings_are_closed(self) -> OriginalReconciliationScopePlan:
        category_codes = [item.code for item in self.categories]
        if len(category_codes) != len(set(category_codes)):
            raise ValueError("scope category codes must be unique")
        coordinates = [(item.sheet_name, item.cell) for item in self.cells]
        if len(coordinates) != len(set(coordinates)):
            raise ValueError("scope cell coordinates must be unique")
        unknown = {item.category_code for item in self.cells} - set(category_codes)
        if unknown:
            raise ValueError("every mapped cell must reference a scope category")
        return self


class OriginalReconciliationImportPlan(_FrozenModel):
    schema_version: Literal["ledgerbridge.original-reconciliation-import-plan.v1"]
    mapping_version: str = Field(pattern=_SAFE_VERSION.pattern)
    source_description: str = Field(min_length=1, max_length=390)
    evidence_filename: str = Field(
        pattern=r"^[A-Za-z0-9][A-Za-z0-9._-]{0,199}$",
    )
    entity: ImportEntity
    scopes: tuple[OriginalReconciliationScopePlan, ...] = Field(min_length=1)

    @model_validator(mode="after")
    def scopes_are_unique(self) -> OriginalReconciliationImportPlan:
        refs = [item.business_unit.ref for item in self.scopes]
        identities = [item.business_unit.business_unit_ref for item in self.scopes]
        if len(refs) != len(set(refs)) or len(identities) != len(set(identities)):
            raise ValueError("business-unit scopes must be unique")
        coordinates = [
            (cell.sheet_name, cell.cell) for scope in self.scopes for cell in scope.cells
        ]
        if len(coordinates) != len(set(coordinates)):
            raise ValueError("a workbook cell can belong to only one business-unit scope")
        return self


def build_original_reconciliation_manifests(
    workbook_path: Path,
    plan: OriginalReconciliationImportPlan,
) -> tuple[SourceManifest, ...]:
    """Return one deterministic controlled-import manifest per business unit."""

    workbook_path = workbook_path.resolve()
    if not workbook_path.is_file():
        raise OriginalReconciliationImportError("workbook path must be a regular file")
    workbook_bytes = workbook_path.read_bytes()
    if len(workbook_bytes) > 134_217_728:
        raise OriginalReconciliationImportError("workbook exceeds the controlled-import limit")
    workbook_sha256 = hashlib.sha256(workbook_bytes).hexdigest()
    generated_at = datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=UTC)

    formulas = load_workbook(workbook_path, read_only=True, data_only=False)
    values = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return tuple(
            _build_scope_manifest(
                formulas=formulas,
                values=values,
                workbook_sha256=workbook_sha256,
                workbook_size=len(workbook_bytes),
                generated_at=generated_at,
                plan=plan,
                scope=scope,
            )
            for scope in plan.scopes
        )
    finally:
        formulas.close()
        values.close()


def _build_scope_manifest(
    *,
    formulas: object,
    values: object,
    workbook_sha256: str,
    workbook_size: int,
    generated_at: datetime,
    plan: OriginalReconciliationImportPlan,
    scope: OriginalReconciliationScopePlan,
) -> SourceManifest:
    identity = (
        f"{workbook_sha256}:{plan.mapping_version}:{plan.entity.entity_ref}:"
        f"{scope.business_unit.ref}"
    )
    evidence_ref = uuid5(_IMPORT_NAMESPACE, identity + ":evidence")
    evidence = SourceEvidence(
        evidence_ref=evidence_ref,
        source_file=plan.evidence_filename,
        display_name=plan.evidence_filename,
        declared_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        plaintext_sha256=workbook_sha256,
        plaintext_size=workbook_size,
    )
    candidates = tuple(
        _build_candidate(
            formulas=formulas,
            values=values,
            identity=identity,
            evidence_ref=evidence_ref,
            mapping=mapping,
        )
        for mapping in scope.cells
    )
    return SourceManifest(
        schema_version="ledgerbridge.controlled-review-source.v1",
        batch_ref=uuid5(_IMPORT_NAMESPACE, identity + ":batch"),
        generated_at=generated_at,
        source_description=f"{plan.source_description} [{plan.mapping_version}]",
        entity=plan.entity,
        business_unit=scope.business_unit,
        categories=scope.categories,
        evidence=(evidence,),
        candidates=candidates,
    )


def _build_candidate(
    *,
    formulas: object,
    values: object,
    identity: str,
    evidence_ref: UUID,
    mapping: OriginalReconciliationCellMapping,
) -> ImportCandidate:
    if mapping.sheet_name not in formulas.sheetnames or mapping.sheet_name not in values.sheetnames:  # type: ignore[attr-defined]
        raise OriginalReconciliationImportError(f"mapped sheet is missing: {mapping.sheet_name}")
    formula_cell = formulas[mapping.sheet_name][mapping.cell]  # type: ignore[index]
    if formula_cell.data_type == "f" or (
        isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
    ):
        raise OriginalReconciliationImportError(
            f"mapped cell is a formula and cannot become a business item: "
            f"{mapping.sheet_name}!{mapping.cell}"
        )
    value = values[mapping.sheet_name][mapping.cell].value  # type: ignore[index]
    amount_minor = _money_minor(value, mapping.sheet_name, mapping.cell)
    amount_minor *= mapping.sign_multiplier
    coordinate = f"{mapping.sheet_name}!{mapping.cell}"
    candidate_identity = f"{identity}:{coordinate}"
    return ImportCandidate(
        candidate_ref=uuid5(_IMPORT_NAMESPACE, candidate_identity + ":candidate"),
        operation_id=uuid5(_IMPORT_NAMESPACE, candidate_identity + ":operation"),
        ingest_channel="CONTROLLED_UPLOAD",
        source_system="original_reconciliation_xlsx",
        source_event_ref=uuid5(_IMPORT_NAMESPACE, candidate_identity + ":source-event"),
        display_label=coordinate,
        category_code=mapping.category_code,
        amount_minor=amount_minor,
        accounting_month=mapping.accounting_month,
        summary=f"{mapping.description} | 原表 {coordinate}",
        confidence_basis_points=10_000,
        evidence_refs=(evidence_ref,),
    )


def _money_minor(value: object, sheet_name: str, cell: str) -> int:
    if isinstance(value, bool) or value is None:
        raise OriginalReconciliationImportError(
            f"mapped cell is not a numeric amount: {sheet_name}!{cell}"
        )
    if isinstance(value, float) and not math.isfinite(value):
        raise OriginalReconciliationImportError(
            f"mapped cell is not a finite amount: {sheet_name}!{cell}"
        )
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise OriginalReconciliationImportError(
            f"mapped cell is not a numeric amount: {sheet_name}!{cell}"
        ) from None
    scaled = amount * 100
    if not scaled.is_finite() or scaled != scaled.to_integral_value():
        raise OriginalReconciliationImportError(
            f"mapped cell has more than two decimal places: {sheet_name}!{cell}"
        )
    result = int(scaled)
    if abs(result) > JSON_SAFE_INTEGER:
        raise OriginalReconciliationImportError(
            f"mapped cell exceeds the supported money range: {sheet_name}!{cell}"
        )
    return result
